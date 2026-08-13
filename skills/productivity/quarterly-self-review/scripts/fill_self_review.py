# -*- coding: utf-8 -*-
"""
季度绩效考核指标设定与自评填写工具
支持两种模式：
  --mode=indicators   季初设定考核维度和工作目标（默认）
  --mode=completion   季末填写完成情况

使用方法：
  1. 改 YEAR 和 QUARTER 变量
  2. 如需设定指标（季初）：python fill_self_review.py --mode=indicators
  3. 如需填写完成情况（季末）：python fill_self_review.py --mode=completion
"""
import openpyxl
import os
import re
import sys
from openpyxl.styles import Alignment, Font

# ==================== 配置项 ====================
YEAR = 2026
QUARTER = "三"  # 一/二/三/四
WEEK_RANGE = (26, 39)  # 该季度覆盖的周次范围（三季度一般26-39周）
# ===============================================

WEEKLY_DIR = r"C:\Users\<user>\Desktop\05_周报述职\个人周报\定"
DUBAN_FILE = r"D:\<知识库>\集团督办工作-个人工作任务\个人督办任务清单.md"
INDICATORS_FILE = rf"C:\Users\<user>\Desktop\05_周报述职\季度绩效考核\{YEAR}年{QUARTER}季度绩效方案指标设定.xlsx"
COMPLETION_FILE = rf"C:\Users\<user>\Desktop\05_周报述职\季度绩效考核\{YEAR}年{QUARTER}季度自评明细.xlsx"


def load_weekly_reports(weekly_dir, week_range):
    """读取指定周次范围内的周报文件"""
    reports = {}
    if not os.path.exists(weekly_dir):
        print(f"⚠ 周报目录不存在: {weekly_dir}")
        return reports

    for fname in sorted(os.listdir(weekly_dir)):
        if not fname.endswith('.xlsx'):
            continue
        m = re.search(r'w(\d+)', fname, re.IGNORECASE)
        if m:
            week_num = int(m.group(1))
            if week_range[0] <= week_num <= week_range[1]:
                fpath = os.path.join(weekly_dir, fname)
                try:
                    wb = openpyxl.load_workbook(fpath, data_only=True)
                    ws = wb.active
                    rows = []
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                        rows.append([v for v in row])
                    reports[week_num] = rows
                    print(f"  ✓ 读取第{week_num}周: {fname}")
                except Exception as e:
                    print(f"  ✗ 读取失败 {fname}: {e}")
    return reports


def extract_summary(reports):
    """从周报中提取本周总结文本"""
    content_parts = []
    for week_num in sorted(reports.keys()):
        rows = reports[week_num]
        summary = []
        in_summary = False
        for row in rows:
            if not row or not row[0]:
                continue
            text = str(row[0])
            if "本周总结" in text:
                in_summary = True
                continue
            if "下周工作重点" in text:
                break
            if in_summary and text.strip():
                summary.append(text.strip())
        if summary:
            content_parts.append(f"【第{week_num}周】\n" + "\n".join(summary))
    return "\n\n".join(content_parts)


def mode_indicators():
    """季初：设定绩效指标"""
    if not os.path.exists(INDICATORS_FILE):
        print(f"⚠ 指标设定文件不存在: {INDICATORS_FILE}")
        print("  请先确认文件路径")
        return

    wb = openpyxl.load_workbook(INDICATORS_FILE)
    ws = wb.active

    print(f"\n=== {YEAR}年{QUARTER}季度绩效方案指标设定 ===")
    print("请在下方为每个考核维度填写工作目标")
    print("（多行以 \\n 分隔，输入 END 结束）\n")

    for row_num in range(2, 7):
        task = ws.cell(row=row_num, column=1).value or ""
        print(f"【行{row_num}】{task}")
        print(f"  当前工作目标：{ws.cell(row=row_num, column=6).value or '(空)'}")
        print(f"  当前权重：{ws.cell(row=row_num, column=5).value or 20}")
        ans = input("  修改？(y/n) ").strip().lower()
        if ans == 'y':
            print("  输入新名称（直接回车保留）:")
            name = input().strip()
            if name:
                ws.cell(row=row_num, column=1).value = name

            print("  输入新权重（直接回车保留当前值）:")
            w = input().strip()
            if w:
                ws.cell(row=row_num, column=5).value = int(w)

            print("  输入工作目标（每行一条，输入 END 结束）:")
            lines = []
            while True:
                line = input()
                if line.strip().upper() == "END":
                    break
                lines.append(line)
            if lines:
                content = "\n".join(lines)
                cell = ws.cell(row=row_num, column=6)
                cell.value = content
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.font = Font(name='Calibri', size=11)
        print()

    wb.save(INDICATORS_FILE)
    print(f"✅ 已保存到: {INDICATORS_FILE}")


def mode_completion():
    """季末：填写完成情况"""
    if not os.path.exists(COMPLETION_FILE):
        print(f"⚠ 自评明细文件不存在: {COMPLETION_FILE}")
        return

    # 1. 读取周报
    print("📖 正在读取周报...")
    reports = load_weekly_reports(WEEKLY_DIR, WEEK_RANGE)
    extracted = extract_summary(reports)
    print(f"\n📝 共读取 {len(reports)} 份周报，提取 {len(extracted)} 字符\n")

    # 2. 读取督办清单（如果有）
    duban_tasks = []
    if os.path.exists(DUBAN_FILE):
        with open(DUBAN_FILE, 'r', encoding='utf-8') as f:
            duban_content = f.read()
        # 提取进行中的督办任务
        task_pattern = re.findall(r'\|\s*(\d+)\s*\|(.+?)\|(.+?)\|(.+?)\|', duban_content)
        for num, content, deadline, status in task_pattern:
            if "已完成" not in status and "已办结" not in status:
                duban_tasks.append(f"  任务{num.strip()}: {content.strip()}")
        if duban_tasks:
            print(f"📋 督办任务（进行中）:")
            for t in duban_tasks:
                print(t)
        print()

    # 3. 逐行填写完成情况
    wb = openpyxl.load_workbook(COMPLETION_FILE)
    ws = wb.active

    print("请在下方填写每个维度的完成情况：")
    print("（多行以 \\n 分隔，输入 END 结束）\n")

    for row_num in range(3, 8):
        task = ws.cell(row=row_num, column=2).value or ""
        goal = ws.cell(row=row_num, column=7).value or ""
        print(f"【行{row_num}】{task}")
        print(f"  工作目标：{str(goal)[:80]}...")
        lines = []
        while True:
            line = input()
            if line.strip().upper() == "END":
                break
            lines.append(line)
        content = "\n".join(lines)
        if content.strip():
            cell = ws.cell(row=row_num, column=8)
            cell.value = content
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.font = Font(name='Calibri', size=11)
        print()

    wb.save(COMPLETION_FILE)
    print(f"✅ 已保存到: {COMPLETION_FILE}")


def main():
    mode = "indicators"
    if len(sys.argv) > 1:
        if sys.argv[1] == "--mode=completion":
            mode = "completion"

    print(f"=== {YEAR}年{QUARTER}季度考核 ===")
    print(f"模式: {'📋 指标设定' if mode == 'indicators' else '📝 完成情况填写'}")

    if mode == "indicators":
        mode_indicators()
    else:
        mode_completion()


if __name__ == "__main__":
    main()
