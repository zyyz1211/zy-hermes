# -*- coding: utf-8 -*-
"""Apply parent company mapping to the new file."""
import openpyxl
import re

input_path = r"C:\Users\<user>\Desktop\<结算工作>\zy改\外部业主_完工未结算_<日期>.xlsx"
output_path = r"C:\Users\<user>\Desktop\<结算工作>\zy改\外部业主_完工未结算_<日期>_带母公司.xlsx"

def match_pattern(owner, patterns):
    for p in patterns:
        if re.search(p, owner):
            return True
    return False

def get_parent(owner):
    if not owner or owner == '-':
        return ''
    
    # --- 1. 某工程集团 ---
    cncec = [r'^中国天辰工程有限公司$', r'^天辰科技园开发（天津）有限公司$', r'^福建天辰耀隆新材料有限公司$',
             r'^中化学天辰（泉州）新材料有限公司$', r'^赛鼎工程有限公司$', r'^某工程集团赛鼎宁波工程有限公司$',
             r'^中化学赛鼎焦化（山西）工程科技有限公司$', r'^中化学赛鼎科创产业发展有限公司$',
             r'^东华工程科技股份有限公司$', r'^中国五环工程有限公司$', r'^武汉天元工程有限责任公司$',
             r'^中国成达工程有限公司$', r'^中化学国际工程有限公司$', r'^中化学交通建设', r'^中化学南方建设投资有限公司$',
             r'^中化学城市投资有限公司$', r'^中化学建设投资集团', r'^中化学生态水利建设有限公司$',
             r'^中化学西部新材料', r'^中化建工程集团北京建设投资有限公司$', r'^中化建工程集团南方建设投资有限公司$',
             r'^某工程集团工程第', r'^中化二建集团有限公司', r'^中化二建集团有限公司第一分公司$', r'^中化二建集团有限公司西南分公司$',
             r'^中化三建', r'^中化六建', r'^中化七建', r'^中化十一建', r'^中化十四建',
             r'^中化学(天津)新材料科技有限公司$',
             r'^中化学五环祥云', r'^中化学华陆新材料有限公司$', r'^某工程集团（股份）有限公司$',
             r'^某工程集团图拉分公司$', r'^华陆工程科技有限责任公司$',
             r'^中化学城投绿景']
    if match_pattern(owner, cncec): return '某工程集团'
    if owner.startswith('中化学') and not any(x in owner for x in ['重庆','泉州','蓝天','蓝星','连云港','东大','高性能','弘润']): return '某工程集团'
    if '天辰科技园' in owner: return '某工程集团'
    
    # --- 2-5. 三桶油 ---
    if match_pattern(owner, [r'^中国石化', r'^中石化', r'^某石化央企股份有限公司']): return '某石化央企'
    if match_pattern(owner, [r'^中国石油天然气', r'^中石油', r'^中国石油集团']) and '某石化央企' not in owner: return '中国石油天然气集团有限公司'
    if match_pattern(owner, [r'^中海石油', r'^中海油', r'^中海福建天然气', r'^中海壳牌', r'^中海沥青', r'^海洋石油富岛']): return '中国海洋石油集团有限公司'
    
    # --- 中化 ---
    if match_pattern(owner, [r'^中化泉州', r'^中化重庆', r'^中化连云港', r'^中化蓝天', r'^中化蓝星', r'^中化东大', r'^中化高性能', r'^中化弘润']): return '中国中化控股有限责任公司'
    # 注意：中化二建、中化三建、中化六建等是 某工程集团 的子公司，不是中化的
    if re.match(r'^中化(?!学)(?!二建)(?!三建)(?!六建)(?!建)', owner): return '中国中化控股有限责任公司'
    
    # --- 6-10. 能源集团 ---
    if match_pattern(owner, [r'^神华', r'^国家能源', r'^国能', r'^中国神华', r'^国电', r'^宁夏煤业']): return '国家能源投资集团有限责任公司'
    # 中煤科工(中国煤炭科工集团)不同于中煤能源集团，先排除
    if match_pattern(owner, [r'^中煤科工']): return '中国煤炭科工集团有限公司'
    if match_pattern(owner, [r'^中煤']): return '中国中煤能源集团有限公司'
    if match_pattern(owner, [r'^陕煤集团', r'^陕煤化', r'^陕西煤业']): return '陕西煤业化工集团有限责任公司'
    if match_pattern(owner, [r'^陕西延长', r'^延长石油']): return '陕西延长石油（集团）有限责任公司'
    if match_pattern(owner, [r'^兖矿', r'^山能', r'^山东能源']): return '山东能源集团有限公司'
    
    # --- 11-20. 电力/工业央企 ---
    if match_pattern(owner, [r'^华能']): return '中国华能集团有限公司'
    if match_pattern(owner, [r'^华电']): return '中国华电集团有限公司'
    if match_pattern(owner, [r'^某央企电力集团', r'^中电投']): return '国家电力投资集团有限公司'
    if match_pattern(owner, [r'^中国某发电集团', r'^某发电集团']): return '中国某发电集团集团有限公司'
    if match_pattern(owner, [r'^中国建材', r'^中建材', r'^中材']): return '中国建材集团有限公司'
    if match_pattern(owner, [r'^中国铝业', r'^中铝']): return '中国铝业集团有限公司'
    if match_pattern(owner, [r'^宝武', r'^宝钢', r'^武钢']): return '中国宝武钢铁集团有限公司'
    if match_pattern(owner, [r'^华润', r'^华润燃气']): return '中国华润有限公司'
    if match_pattern(owner, [r'^中盐']): return '中国盐业集团有限公司'
    if match_pattern(owner, [r'^中国船舶', r'^中船']): return '中国船舶集团有限公司'
    if match_pattern(owner, [r'^浙江省能源', r'^浙能']): return '浙江省能源集团有限公司'
    if match_pattern(owner, [r'^伊犁新天']): return '浙江省能源集团有限公司'
    if match_pattern(owner, [r'^中国兵器', r'^中兵', r'^北方工业']): return '中国某央企军工集团集团有限公司'
    if match_pattern(owner, [r'^阿克苏华锦']): return '中国某央企军工集团集团有限公司'
    if match_pattern(owner, [r'^航天', r'^中国航天']): return '中国航天科技集团有限公司'
    if match_pattern(owner, [r'^中国核工业', r'^中核']): return '中国核工业集团有限公司'
    if match_pattern(owner, [r'^中冶']): return '中国冶金科工集团有限公司'
    if match_pattern(owner, [r'^中国电建', r'^中电建', r'^中国水利水电']): return '中国电力建设集团有限公司'
    if match_pattern(owner, [r'^中国能建', r'^中能建']): return '中国能源建设集团有限公司'
    if match_pattern(owner, [r'^中交', r'^中国交建']): return '中国交通建设集团有限公司'
    if match_pattern(owner, [r'^中国铁建']): return '中国铁建股份有限公司'
    if match_pattern(owner, [r'^中铁\d{2,3}']): return '中国中铁股份有限公司'
    if match_pattern(owner, [r'^中国建筑', r'^中建']): return '中国建筑集团有限公司'
    if match_pattern(owner, [r'^中信']): return '中国某综合集团有限公司'
    if match_pattern(owner, [r'^中国航空', r'^中国航油', r'^中国航发']): return '中国航空工业集团有限公司'
    if match_pattern(owner, [r'^中国广核', r'^中广核']): return '中国广核集团有限公司'
    if match_pattern(owner, [r'^中国节能', r'^中节能']): return '中国节能环保集团有限公司'
    if match_pattern(owner, [r'^中国有色', r'^中色']): return '中国有色矿业集团有限公司'
    if match_pattern(owner, [r'^中国保利', r'^保利']): return '中国保利集团有限公司'
    if match_pattern(owner, [r'^招商局', r'^招商']): return '招商局集团有限公司'
    if match_pattern(owner, [r'^国网']): return '国家电网有限公司'
    if match_pattern(owner, [r'^中国电子', r'^中电科']): return '中国电子信息产业集团有限公司'
    
    # --- 化工企业集团 ---
    if match_pattern(owner, [r'^万华化学', r'^万华科化']): return '万华化学集团股份有限公司'
    if match_pattern(owner, [r'^湖北宜化']): return '湖北宜化集团有限责任公司'
    if match_pattern(owner, [r'^湖北兴发', r'^湖北兴瑞', r'^湖北兴力', r'^湖北泰盛', r'^湖北兴福',
                             r'^襄阳兴发', r'^内蒙古兴发', r'^宜都兴发', r'^湖北兴宏',
                             r'^湖北吉星', r'^湖北瑞佳', r'^湖北兴磷', r'^湖北兴晨']): return '宜昌兴发集团有限责任公司'
    if match_pattern(owner, [r'^荣盛', r'^浙江荣盛']): return '浙江荣盛控股集团有限公司'
    if match_pattern(owner, [r'^恒力']): return '恒力集团有限公司'
    if match_pattern(owner, [r'^盛虹']): return '盛虹控股集团有限公司'
    if match_pattern(owner, [r'^华鲁']): return '华鲁控股集团有限公司'
    if match_pattern(owner, [r'^东明']): return '山东某石化企业集团有限公司'
    if match_pattern(owner, [r'^京博']): return '京博控股集团有限公司'
    if match_pattern(owner, [r'^贵州磷化', r'^瓮福', r'^贵州开磷', r'^贵阳开磷']): return '贵州磷化（集团）有限责任公司'
    if match_pattern(owner, [r'^云南云天化', r'^云南天安', r'^云南水富云天化', r'^云南三环', r'^重庆云天化', r'^云南磷化']): return '云天化集团有限责任公司'
    if match_pattern(owner, [r'^河南能源', r'^永煤', r'^义煤', r'^鹤煤']): return '河南能源集团有限公司'
    if match_pattern(owner, [r'^晋能', r'^晋控', r'^晋煤']): return '晋能控股集团有限公司'
    if match_pattern(owner, [r'^潞安']): return '潞安化工集团有限公司'
    if match_pattern(owner, [r'^新疆中泰', r'^中泰化学']): return '新疆中泰（集团）有限责任公司'
    if match_pattern(owner, [r'^新疆天业', r'^天业']): return '新疆天业（集团）有限公司'
    if match_pattern(owner, [r'^特变电工']): return '特变电工股份有限公司'
    if match_pattern(owner, [r'^金川集团']): return '金川集团股份有限公司'
    if match_pattern(owner, [r'^旭阳']): return '旭阳集团有限公司'
    if match_pattern(owner, [r'^美锦']): return '美锦能源集团有限公司'
    if match_pattern(owner, [r'^广汇']): return '广汇能源股份有限公司'
    if match_pattern(owner, [r'^淮北矿业']): return '淮北矿业（集团）有限责任公司'
    if match_pattern(owner, [r'^华阳新材料']): return '华阳新材料科技集团有限公司'
    if match_pattern(owner, [r'^宜宾天原']): return '宜宾天原集团股份有限公司'
    if match_pattern(owner, [r'^陕西投资']): return '陕西投资集团有限公司'
    if match_pattern(owner, [r'^陕西金泰氯碱']): return '陕西投资集团有限公司'
    if match_pattern(owner, [r'^陕西榆林能源', r'^陕西榆能']): return '陕西榆林能源集团有限公司'
    if match_pattern(owner, [r'^陕西精益']): return '陕西榆林能源集团有限公司'
    if match_pattern(owner, [r'^陕西有色金属']): return '陕西有色金属控股集团有限责任公司'
    if match_pattern(owner, [r'^包头钢铁', r'^包钢']): return '包头钢铁（集团）有限责任公司'
    if match_pattern(owner, [r'^甘肃能源化工', r'^甘肃能化']): return '甘肃能源化工投资集团有限公司'
    if match_pattern(owner, [r'^福建能源石化', r'^福建省能源石化']): return '某能源集团有限责任公司'
    if match_pattern(owner, [r'^上海华谊']): return '上海华谊集团股份有限公司'
    if match_pattern(owner, [r'^天津渤海化工']): return '天津渤海化工集团有限责任公司'
    if match_pattern(owner, [r'^贵州能源']): return '贵州能源集团有限公司'
    if match_pattern(owner, [r'^陕西省燃气', r'^陕西燃气']): return '陕西燃气集团有限公司'
    if match_pattern(owner, [r'^江苏省国信']): return '江苏省国信集团有限公司'
    if match_pattern(owner, [r'^旭阳']): return '旭阳集团有限公司'
    if match_pattern(owner, [r'^合盛硅业']): return '合盛硅业股份有限公司'
    if match_pattern(owner, [r'^卫星化学', r'^浙江卫星']): return '卫星化学股份有限公司'
    if match_pattern(owner, [r'^广西华谊']): return '上海华谊集团股份有限公司'
    if match_pattern(owner, [r'^山西亚鑫']): return '山西亚鑫能源集团有限公司'
    if match_pattern(owner, [r'^山西美锦', r'^美锦华盛']): return '山西美锦能源股份有限公司'
    if match_pattern(owner, [r'^天津渤化永利', r'^天津长芦汉沽']): return '天津渤海化工集团有限责任公司'
    if match_pattern(owner, [r'^贵州瓮福开磷', r'^贵州瓮福江山']): return '贵州磷化（集团）有限责任公司'
    if match_pattern(owner, [r'^金风绿能']): return '金风科技股份有限公司'
    if match_pattern(owner, [r'^唐山中浩']): return '开滦（集团）有限责任公司'
    if match_pattern(owner, [r'^福建中燃']): return '中国燃气控股有限公司'
    if match_pattern(owner, [r'^河南神马尼龙']): return '中国平煤神马控股集团有限公司'
    if match_pattern(owner, [r'^利华益']): return '利华益集团股份有限公司'
    if match_pattern(owner, [r'^海科新源']): return '山东海科控股有限公司'
    if match_pattern(owner, [r'^营口建发盛海']): return '厦门建发集团有限公司'
    if match_pattern(owner, [r'^信义硅业']): return '信义光能控股有限公司'
    if match_pattern(owner, [r'^海东红狮']): return '红狮控股集团有限公司'
    if match_pattern(owner, [r'^江西晶昊']): return '江西盐业集团有限公司'
    if match_pattern(owner, [r'^铜陵悦江首创']): return '北京首创生态环保集团股份有限公司'
    if match_pattern(owner, [r'^北京金隅北水']): return '北京金隅集团股份有限公司'
    if match_pattern(owner, [r'^连云港圣奥']): return '圣奥化学科技有限公司'
    if match_pattern(owner, [r'^苏利（宁夏）']): return '苏利农业科技股份有限公司'
    if match_pattern(owner, [r'^四川美丰']): return '四川美丰化工股份有限公司'
    if match_pattern(owner, [r'^弘元能源']): return '弘元绿色能源股份有限公司'
    if match_pattern(owner, [r'^山西兰花']): return '山西兰花煤炭实业集团有限公司'
    if match_pattern(owner, [r'^四川泸天化']): return '四川泸天化股份有限公司'
    if match_pattern(owner, [r'^山西梗阳']): return '山西梗阳投资集团有限公司'
    if match_pattern(owner, [r'^山西闽光']): return '山西闽光新能源科技股份有限公司'
    if match_pattern(owner, [r'^江苏德邦兴华']): return '江苏德邦化学工业集团有限公司'
    if match_pattern(owner, [r'^山西茂胜']): return '山西茂胜煤化集团有限公司'
    if match_pattern(owner, [r'^山西骏捷']): return '山西骏捷新材料科技有限公司'
    if match_pattern(owner, [r'^呼伦贝尔金新']): return '云南云天化股份有限公司'
    if match_pattern(owner, [r'^四川天华时代']): return '天华新能'
    if match_pattern(owner, [r'^宁夏鲲鹏']): return '宁夏鲲鹏清洁能源有限公司'
    if match_pattern(owner, [r'^江苏索普聚酯']): return '江苏索普（集团）有限公司'
    if match_pattern(owner, [r'^唐山三友硅业']): return '唐山三友化工股份有限公司'
    if match_pattern(owner, [r'^哈尔滨锅炉厂']): return '哈尔滨电气集团有限公司'
    if match_pattern(owner, [r'^新疆圣雄氯碱']): return '新疆圣雄能源开发有限公司'
    if match_pattern(owner, [r'^优利德（湖北）']): return '优利德集团有限公司'
    if match_pattern(owner, [r'^新浦化学']): return '新浦化学（泰兴）有限公司'
    if match_pattern(owner, [r'^重庆市中润化学']): return '重庆市中润化学有限公司'
    if match_pattern(owner, [r'^巴斯夫一体化基地']): return '巴斯夫集团（BASF SE）'
    if match_pattern(owner, [r'^惠州宇新']): return '宇新股份'
    if match_pattern(owner, [r'^湖北昭君古镇']): return '宜昌兴发集团有限责任公司'
    if match_pattern(owner, [r'^乌兰察布市旭峰']): return '乌兰察布市旭峰合源化工有限公司'
    if match_pattern(owner, [r'^甘肃创动航醇']): return '甘肃创动航醇新能源科技有限公司'
    if match_pattern(owner, [r'^湖北黄冈华兴冶金']): return '湖北黄冈华兴冶金窑炉有限公司'
    if match_pattern(owner, [r'^内蒙古西部天然气']): return '内蒙古西部天然气股份有限公司'
    if match_pattern(owner, [r'^国家管网集团']): return '国家石油天然气管网集团有限公司'
    if match_pattern(owner, [r'^辽宁方大工程设计']): return '辽宁方大工程设计有限公司'
    if match_pattern(owner, [r'^中国平煤神马', r'^平煤神马']): return '中国平煤神马控股集团有限公司'
    if match_pattern(owner, [r'^内蒙古伊泰']): return '内蒙古伊泰集团有限公司'
    if match_pattern(owner, [r'^内蒙古伊东']): return '内蒙古伊东资源集团股份有限公司'
    if match_pattern(owner, [r'^内蒙古博源', r'^博源']): return '内蒙古博源控股集团有限公司'
    if match_pattern(owner, [r'^久泰']): return '久泰集团有限公司'
    if match_pattern(owner, [r'^内蒙古君正']): return '内蒙古君正能源化工集团股份有限公司'
    if match_pattern(owner, [r'^内蒙古汇能']): return '内蒙古汇能煤电集团有限公司'
    if match_pattern(owner, [r'^内蒙古双欣']): return '内蒙古双欣能源化工有限公司'
    if match_pattern(owner, [r'^内蒙古广纳']): return '内蒙古广纳煤业（集团）有限责任公司'
    if match_pattern(owner, [r'^内蒙古鄂尔多斯投资']): return '内蒙古鄂尔多斯投资控股集团有限公司'
    if match_pattern(owner, [r'^宁夏宝丰']): return '宁夏宝丰能源集团股份有限公司'
    if match_pattern(owner, [r'^陕西黑猫']): return '陕西黑猫焦化股份有限公司'
    if match_pattern(owner, [r'^成都云图']): return '成都云图控股股份有限公司'
    if match_pattern(owner, [r'^中国北方工业']): return '中国北方工业有限公司'
    if match_pattern(owner, [r'^中国天楹']): return '中国天楹股份有限公司'
    if match_pattern(owner, [r'^云南能投']): return '云南省能源投资集团有限公司'
    if match_pattern(owner, [r'^新疆金风', r'^金风科技']): return '金风科技股份有限公司'
    if match_pattern(owner, [r'^湖南省轻工']): return '湖南省轻工盐业集团有限公司'
    if match_pattern(owner, [r'^永荣']): return '永荣控股集团有限公司'
    if match_pattern(owner, [r'^巨正源']): return '巨正源股份有限公司'
    if match_pattern(owner, [r'^中国燃气']): return '中国燃气控股有限公司'
    if match_pattern(owner, [r'^新疆生产建设']): return '新疆生产建设兵团'
    if match_pattern(owner, [r'^陕西省水务']): return '陕西省水务集团有限公司'
    if match_pattern(owner, [r'^陕西未来能源']): return '陕西未来能源化工有限公司'
    if match_pattern(owner, [r'^陕西榆林能源', r'^陕西榆能']): return '陕西榆林能源集团有限公司'
    if match_pattern(owner, [r'^河南心连心']): return '河南心连心化学工业集团股份有限公司'
    if match_pattern(owner, [r'^安徽海螺']): return '安徽海螺集团有限责任公司'
    if match_pattern(owner, [r'^安徽晋煤中能']): return '安徽晋煤中能化工股份有限公司'
    if match_pattern(owner, [r'^安徽皖维']): return '安徽皖维集团有限责任公司'
    if match_pattern(owner, [r'^开滦']): return '开滦（集团）有限责任公司'
    if match_pattern(owner, [r'^新奥']): return '新奥天然气股份有限公司'
    if match_pattern(owner, [r'^济民可信']): return '济民可信集团有限公司'
    if match_pattern(owner, [r'^亨通']): return '亨通集团有限公司'
    if match_pattern(owner, [r'^黑龙江新产业投资']): return '黑龙江省新产业投资集团有限公司'
    if match_pattern(owner, [r'^东岳']): return '东岳集团有限公司'
    if match_pattern(owner, [r'^湖北和远']): return '湖北和远气体股份有限公司'
    if match_pattern(owner, [r'^北京能源']): return '北京能源集团有限责任公司'
    if match_pattern(owner, [r'^西藏矿业']): return '西藏矿业发展股份有限公司'
    if match_pattern(owner, [r'^靖远煤业']): return '靖远煤业集团有限责任公司'
    if match_pattern(owner, [r'^西安凯立']): return '西安凯立新材料股份有限公司'
    if match_pattern(owner, [r'^石大胜华', r'^石大胜达']): return '石大胜华新材料集团股份有限公司'
    if match_pattern(owner, [r'^烟台泰和']): return '烟台泰和新材料股份有限公司'
    if match_pattern(owner, [r'^中圣']): return '中圣集团'
    if match_pattern(owner, [r'^鞍钢', r'^攀钢']): return '鞍钢集团有限公司'
    if match_pattern(owner, [r'^青海汇信']): return '青海汇信资产管理有限责任公司'
    
    # --- Government / Public ---
    gov = [r'人民政府$', r'管理委员会$', r'局$', r'委员会$', r'办公室$', r'部$', r'^中华人民共和国', r'^解放军', r'^中国科学院', r'^中国科大']
    if match_pattern(owner, gov):
        m = re.match(r'(.+?)人民政府$', owner)
        if m: return f'{m.group(1)}人民政府'
        m = re.match(r'(.+?)(?:住房和城乡建设局|交通运输局|教育局|卫生健康委员会|财政局|发展和改革委员会|水利局|园林绿化局|城市管理局|自然资源和规划局|工业和信息化局|公安局)$', owner)
        if m: return f'{m.group(1)}人民政府'
        return owner
    
    # --- Foreign companies ---
    foreign = {
        'BASF': '巴斯夫集团（BASF SE）',
        'BorsodChem Zrt.': '万华化学集团股份有限公司',
        'SABIC Saudi European Petrochemical Company': '沙特基础工业公司（SABIC）',
        'EuroChem Karatau': '欧洲化学集团（EuroChem Group）',
        'Kazakhstan Petrochemical Industries': '哈萨克斯坦国家石油天然气公司（KMG）',
        'Middle East Kimiaye Pars': '伊朗国家石化公司（NPC）',
        '阿科玛（中国）投资有限公司': '阿科玛集团（Arkema）',
        '科思创聚合物（中国）有限公司': '科思创集团（Covestro）',
        '贺利氏电化（上海）有限公司': '贺利氏集团（Heraeus）',
        '沙特Sabic SADAF烧碱公司': '沙特基础工业公司（SABIC）',
        '丹格特': '丹格特集团（Dangote Group）',
        '联合太阳能': '联合太阳能集团（United Solar）',
        '俄罗斯天然气工业石油公司': '俄罗斯天然气工业股份公司（Gazprom）',
        '阿尔及利亚国家天然气电力集团公司': '阿尔及利亚国家石油公司（Sonatrach）',
        '阿布扎比石油公司': '阿布扎比国家石油公司（ADNOC）',
        'Indorama corporation': 'Indorama Ventures Public Company Limited',
        'PT CHENGTOK LITHIUM INDONESIA': '盛屯矿业集团股份有限公司',
        'PT Indonesia BTR New Energy Material': '贝特瑞新材料集团股份有限公司',
        'PT.OBI NICKEL COBALT': '宁波力勤资源科技股份有限公司',
        '联化科技马来西亚公司': '联化科技股份有限公司',
        '中鼎工程股份有限公司（新加坡分公司）': '中鼎工程股份有限公司（CTCI）',
        '联合技术大宇': '大宇工程建设公司（Daewoo E&C）',
        '乐金化学惠州化工有限公司': 'LG化学（LG Chem）',
        '东邦化学（上海）有限公司': '东邦化学工业株式会社',
        '以化（张家港）国际贸易有限公司': '以色列化工集团（ICL）',
        '中海壳牌石油化工有限公司': '中国海洋石油集团有限公司 / 壳牌集团',
        '中石化英力士苯领': '某石化央企 / Ineos集团',
        '蓝星安迪苏南京有限公司': '中国中化控股有限责任公司',
        '科莱恩华锦催化剂(盘锦)有限公司': '科莱恩集团（Clariant）',
        '迪艾基（广州）气体有限公司': '液化空气集团（Air Liquide）',
        '赢创特种化学（南京）有限公司': '赢创工业集团（Evonik）',
        '空气化工产品（南京）有限公司': '空气化工产品公司（Air Products）',
        '道达尔润滑油（中国）有限公司': '道达尔能源集团（TotalEnergies）',
        '亚什兰化工（南京）有限公司': '亚什兰全球控股公司（Ashland）',
        '西门子数控（南京）有限公司': '西门子股份公司（Siemens）',
        'Ceylon Petroleum Corporation': '锡兰石油公司（CPC）',
        'PowerChina International Group Limited Egypt Branch': '中国电力建设集团有限公司',
        'TENKE FUNGURUME MINING S.A': '中国有色矿业集团有限公司',
        'Tsingshan Mining Development SA': '青山控股集团有限公司',
    }
    for k, v in foreign.items():
        if k in owner or owner == k: return v
    
    # --- Self-parent companies (keep their own name) ---
    self_parents = [
        '金沂蒙集团有限公司', '济南圣泉集团股份有限公司', '开山控股集团股份有限公司',
        '新疆新业国有资产经营(集团)有限责任公司', '新疆雪峰科技（集团）股份有限公司',
        '新疆庆华能源集团有限公司', '新疆美克化工股份有限公司', '湖北三宁化工股份有限公司',
        '广东大鹏液化天然气有限公司', '内蒙古美方煤焦化有限公司',
        '内蒙古恒星化学有限公司', '宁波金海晨光化学股份有限公司',
        '铜陵贝斯美科技有限公司', '山东方明化工股份有限公司',
        '山东蓝湾新材料有限公司', '福建百宏化学有限公司',
        '山西蔺鑫煤焦化有限责任公司', '陕西精益化工有限公司',
        '陕西润中清洁能源有限公司', '陕西渭河彬州化工有限公司',
        '云南珠江实业集团有限公司', '贵州胜泽威化工有限公司',
        '贵州新天鑫化工有限公司', '海南华盛新材料科技有限公司',
        '唐山裕隆新材料科技有限公司', '洪洞县民生垃圾综合处理有限公司',
        '河南碳氢重业材料有限公司', '河南永昌硝基肥有限公司',
        '太原金翰源科技有限公司', '宁夏凯信能源科技有限公司',
        '安徽碳鑫科技有限公司', '新疆泰亨能源化工有限责任公司',
        '新疆兴准能源有限公司', '新疆天达新材料有限公司',
        '新疆国业新材料科技有限公司', '新疆伊力特煤化工有限责任公司',
        '长盛（廊坊）科技有限公司', '河北中科富峰氢能科技有限公司',
        '介休市昌盛煤气化有限公司', '呼伦贝尔东北草丰生物科技有限公司',
        '新疆中能绿源化工有限公司', '安阳瑞美达清洁能源有限公司',
        '上海臻友设备工程技术有限公司', '湖北新源浩科新材料有限公司',
        '新疆锦疆化工股份有限公司', '新疆宣东能源有限公司',
        '新疆天运化工有限公司', '新疆曙光绿华生物科技有限公司',
        '时代思康新材料有限公司', '新能能源有限公司',
        '重庆建峰兴源科技有限公司', '华瀛天然气股份有限公司',
        '山东鲁南园区投资建设有限公司', '新疆至创新材料有限公司',
        '内蒙古华恒能源科技有限公司', '内蒙古卓正煤化工有限公司',
        '内蒙古东日新能源材料有限公司', '内蒙古兴洋科技股份有限公司',
        '内蒙古美邦中科新材料有限公司', '内蒙古中谷矿业有限责任公司',
        '内蒙古中能生物科技有限公司', '内蒙古大地云天化工有限公司',
        '内蒙古瑞志现代煤化工科技有限公司', '内蒙古恒坤化工有限公司',
        '内蒙古东立光伏电子有限公司', '内蒙古东立光伏股份有限公司',
        '兴安盟乌兰泰安能源化工有限责任公司', '内蒙古华景新材料有限责任公司',
        '内蒙古吉源热电有限责任公司', '内蒙古嘉洋科技有限公司',
        '内蒙古圣钒科技新能源有限责任公司', '内蒙古启杭新能源有限公司',
        '内蒙古东源科技集团有限公司', '内蒙古三维新材料有限公司',
        '内蒙古大全新能源有限公司', '内蒙古新特硅材料有限公司',
        '内蒙古永和氟化工有限公司', '内蒙古永太化学有限公司',
        '内蒙古润阳悦达新能源科技有限公司', '内蒙古广聚新材料有限责任公司',
        '内蒙古宝丰煤基新材料有限公司', '湖北新源浩科新材料有限公司',
        '宁夏百川科技有限公司', '宁夏百川新材料有限公司',
        '福建华星石化有限公司', '湖北省黄麦岭控股集团有限公司',
        '宁夏蛋氨酸有限公司', '甘肃耀望化工有限公司',
        '新疆华醇能源有限公司', '新疆励晶煤业有限公司',
        '浙江天硅新材料有限公司', '泰兴金燕化学科技有限公司',
        '无锡翔龙环球科技股份有限公司', '博大东方新型化工（吉林）有限公司',
        '山西冠力法兰股份有限公司', '云南解化清洁能源开发有限公司',
        '杭州和兴碳纤维科技有限公司', '阿拉善盟沪蒙能源实业有限公司',
        '铁岭选矿药剂有限公司', '陕西中鑫万利环保科技有限公司',
        '福建福维新材料有限公司', '山东顺东港务有限公司',
        '贵州溢鑫实业投资有限责任公司', '磊泰科技有限公司',
        '浙江三江思怡新材料有限公司', '大连金重机器集团有限公司',
        '山东天辰新材料科技有限公司', '中天东方氟硅材料有限公司',
        '浙江中天东方氟硅材料股份有限公司', '浙江创世雷博科技有限公司',
        '天辰化工有限公司', '安徽普盛医疗科技有限公司',
        '西藏阿里麻米措矿业开发有限公司', '山西沃能化工科技有限公司',
        '南宁科天水性科技有限责任公司', '营口佳孚石油化工有限公司',
        '埃得新材料有限公司', '鄂尔多斯市双欣化学工业有限责任公司',
        '广西鹏越生态科技有限公司', '陕西渭河重化工有限责任公司',
        '蚌埠工投科技发展集团有限公司', '淮安兴盛建设投资有限公司',
        '鹤壁龙宇新材料有限公司', '湖北姚家港绿色化工投资控股集团有限公司',
        '濮阳朗润新材料有限公司', '山西转型综改示范区合成生物产业投资开发有限公司',
        '陕西榆神能源通达管网有限责任公司', '陕西水务科创产业发展有限公司',
        '陕西龙华集团煤业科技发展有限公司', '西安凯立新材料股份有限公司',
        '沧州旭阳化工有限公司', '胜帮科技股份有限公司',
        '重庆建峰工业集团有限公司', '孝义市盛世富源甲醇制造有限公司',
        '新疆其亚化工有限公司', '新疆至臻化工工程研究中心有限公司',
        '新疆天智辰业化工有限公司', '石河子天域新实化工有限公司',
        '新疆新业能源化工有限责任公司', '山东晋控日月新材料有限公司',
        '兖州煤业榆林能化有限公司', '兖矿新疆煤化工公司', '兖矿鲁南化工有限公司',
        '陕西有色天宏瑞科硅材料有限责任公司', '陕西未来清洁化学品有限公司',
        '陕西金泰氯碱化工有限公司', '陕西金泰氯碱神木化工有限公司',
        '陕西省天然气股份有限公司', '陕西延长中煤榆林能源化工股份有限公司',
        '陕西延长石油延安能源化工有限责任公司', '陕西延长石油榆林煤化有限公司',
        '陕西延长石油（集团）有限责任公司', '神华工程技术有限公司',
        '神华新疆能源有限责任公司', '中煤陕西榆林能源化工有限公司',
        '中煤鄂尔多斯能源化工有限公司', '中煤平朔集团有限公司',
        '中煤内蒙古能源有限公司', '内蒙古包钢钢联股份有限公司',
        '内蒙古包钢低碳产业科技发展有限公司',
        '内蒙古鄂尔多斯电力冶金集团股份有限公司氯碱化工分公司',
        '广东大鹏液化天然气有限公司', '国化融资租赁(天津)有限公司',
        '福建福海创石油化工有限公司', '福建省福化天辰气体有限公司',
        '伊犁新天煤化工有限责任公司', '福建申远新材料有限公司',
        '磴口陕耀氢能新能源有限公司', '内蒙古某发电集团国际克什克腾煤制天然气有限责任公司',
        '山东裕龙石化有限公司', '新疆兵能煤业有限责任公司',
        '新疆新冀能源化工有限公司', '广东省盛元达科技有限公司',
        '金开新疆煤制气有限公司', '安达市天楹新能源有限公司',
        '新疆黑猫煤化工有限公司', '同煤广发化学工业有限公司',
        '应城市新都化工有限责任公司', '湖北新宜化工有限公司',
        '湖北楚星化工股份有限公司', '贵州盘江电投天能焦化有限公司',
        '云南三环中化化肥有限公司', '湖北宜化新能源有限公司',
        '湖北宜化碳一化工有限公司', '湖北宜化精细化工有限公司',
        '湖北宜化磷化工有限公司', '湖北宜化氟化工有限公司',
        '湖北宜化楚星生态科技有限公司', '云南解化清洁能源开发有限公司解化化工分公司',
        '重庆湘渝盐化有限责任公司', '河南晋开集团延化化工有限公司',
        '阿克苏华锦化肥有限责任公司', '湖北双环科技股份有限公司',
        '天津渤化化工发展有限公司', '甘肃启化新材料有限责任公司',
        '山东中燃宝港能源发展有限公司', '天津中圣泰港新能源科技有限公司',
        '昆明云能化工有限公司', '新疆中能万源化工有限公司',
        '陕西榆能能化新材料有限公司', '福建福杭新业科技股份有限公司',
        '上海华谊工业气体有限公司', '新疆山能化工有限公司',
        '宁夏宁东泰和化学科技有限公司', '福建永荣科技有限公司',
        '福建海泉化学有限公司', '河南永银化工实业有限公司',
        '拜城县众泰煤焦化有限公司', '兰州盈德气体有限公司',
        '新疆广汇新能源有限公司', '新疆中泰新材料股份有限公司',
        '贵州美锦华宇新能源有限公司', '山西华阳碳材科技有限公司',
        '新汶矿业集团有限责任公司', '河北开滦航橡新材料有限公司',
        '黑龙江省龙江化工有限公司', '云南能投硅材科技发展有限公司',
        '金风绿能化工（兴安盟）有限公司', '湖北和远新材料有限公司',
        '灵石县中煤九鑫焦化有限责任公司', '北京昊华能源股份有限公司',
        '灵石中煤化工有限责任公司', '山东东岳有机硅材料股份有限公司',
        '江山双阳水有限公司', '宁夏英力特化工股份有限公司',
        '陕煤集团榆林化学宇高新材料有限责任公司', '陕煤集团榆林化学有限公司',
        '陕西榆能化学材料有限公司', '阳煤平原化工有限公司',
        '宜宾天原海丰和泰有限公司', '临涣焦化股份有限公司',
        '安徽华塑股份有限公司', '青海汇信新材料科技有限公司',
        '西藏日喀则扎布耶锂业高科技有限公司', '森特士兴集团股份有限公司合肥分公司',
        '安徽皖维高新材料股份有限公司', '铜川市天然气有限公司',
        '江苏国信液化天然气有限公司', '宜都兴发生态园区开发有限公司',
        '济民可信（高安）清洁能源有限公司', '云图新能源材料（荆州）有限公司',
        '内蒙古亨芯石英有限公司', '河南开祥精细化工有限公司',
        '新疆心连心能源化工有限公司', '湖北瑞佳硅材料有限公司',
        '河南神马氢化学有限责任公司', '河南神马尼龙化工有限责任公司',
        '湖北兴福电子材料股份有限公司', '湖北兴宏矿业有限公司',
        '湖北兴瑞硅材料有限公司', '湖北泰盛化工有限公司',
        '湖北兴力电子材料有限公司', '襄阳兴发化工有限公司',
        '内蒙古兴发科技有限公司', '湖北吉星化工集团有限责任公司',
        '湖北新宜化工有限公司', '湖北泰盛化工有限公司',
        '石大胜达（泉州）有限公司', '宁夏煤业有限责任公司煤制油分公司',
        '安徽海螺材料科技股份有限公司', '凯立新材（彬州）科技有限公司',
        '阿拉尔青松化工有限责任公司', '天津谊德国际贸易有限公司',
        '福建德胜能源有限公司', '宁波大榭化工仓储有限公司',
        '孝义市盛世富源甲醇制造有限公司', '陕西渭河彬州化工有限公司',
        '浙江卓锦环保科技股份有限公司', '河南碳氢重业材料有限公司',
        '山东蓝湾新材料有限公司', '新疆励晶煤业有限公司',
        '浙江天硅新材料有限公司', '泰兴金燕化学科技有限公司',
        '无锡翔龙环球科技股份有限公司', '福建福维新材料有限公司',
        '山东顺东港务有限公司', '贵州溢鑫实业投资有限责任公司',
        '磊泰科技有限公司', '浙江三江思怡新材料有限公司',
        '大连金重机器集团有限公司', '铜陵贝斯美科技有限公司',
        '山东天辰新材料科技有限公司', '中天东方氟硅材料有限公司',
        '浙江中天东方氟硅材料股份有限公司', '浙江创世雷博科技有限公司',
        '天辰化工有限公司', '安徽普盛医疗科技有限公司',
        '海南华盛新材料科技有限公司', '荆门源晗电池材料有限公司',
        '西藏阿里麻米措矿业开发有限公司', '山西沃能化工科技有限公司',
        '山东方明化工股份有限公司', '贵州新天鑫化工有限公司',
        '合肥高新股份有限公司', '广西鹏越生态科技有限公司',
        '蚌埠工投科技发展集团有限公司', '南宁科天水性科技有限责任公司',
        '埃得新材料有限公司', '新疆中能绿源化工有限公司',
        '安阳瑞美达清洁能源有限公司', '上海臻友设备工程技术有限公司',
        '湖北新源浩科新材料有限公司', '福建百宏化学有限公司',
        '昆明胜威化工有限公司', '淮安兴盛建设投资有限公司',
        '鹤壁龙宇新材料有限公司', '湖北姚家港绿色化工投资控股集团有限公司',
        '濮阳朗润新材料有限公司', '新疆锦疆化工股份有限公司',
        '宁夏百川科技有限公司', '宁夏百川新材料有限公司',
        '福建华星石化有限公司', '湖北省黄麦岭控股集团有限公司',
        '宁夏蛋氨酸有限公司', '甘肃耀望化工有限公司',
        '陕西榆神能源通达管网有限责任公司', '陕西水务科创产业发展有限公司',
        '陕西龙华集团煤业科技发展有限公司',
        '山西转型综改示范区合成生物产业投资开发有限公司',
        '鄂尔多斯市双欣化学工业有限责任公司', '内蒙古东源科技集团有限公司',
        '兴安盟博源化学有限公司', '内蒙古伊泰化工有限责任公司',
        '内蒙古伊东集团九鼎化工有限责任公司', '内蒙古博大实地化学有限公司',
        '内蒙古博源银根化工有限公司', '内蒙古博源银根矿业有限责任公司',
        '内蒙古久泰新材料有限公司', '内蒙古久泰新材料科技股份有限公司',
        '内蒙古久泰馨远新材料有限公司', '内蒙古君正化工有限责任公司',
        '内蒙古双欣环保材料股份有限公司', '内蒙古宜化化工有限公司',
        '内蒙古三爱富万豪氟化工有限公司', '内蒙古东景生物环保科技有限公司',
        '胜帮科技股份有限公司', '国化融资租赁(天津)有限公司',
        '新疆新业国有资产经营(集团)有限责任公司',
        '准格尔旗鼎鑫工贸有限责任公司',
        '哈密茂坤能源有限公司',
        '哈密恒有能源化工科技有限公司',
        '上海电气集团国控环球工程有限公司',
    ]
    for p in self_parents:
        if p == owner or p in owner:
            return owner
    
    return ''

# Main processing
wb = openpyxl.load_workbook(input_path, data_only=True)
ws = wb.active

# Insert column I after H (column 8)
ws.insert_cols(9)
ws.cell(row=3, column=9, value='母公司名称')

mapped = 0
unmapped = []
for row in ws.iter_rows(min_row=4, max_col=9, max_row=ws.max_row, values_only=False):
    owner_cell = row[7]
    parent_cell = row[8]
    if owner_cell.value and str(owner_cell.value).strip():
        owner = str(owner_cell.value).strip()
        parent = get_parent(owner)
        parent_cell.value = parent
        if parent:
            mapped += 1
        else:
            if owner not in unmapped:
                unmapped.append(owner)

# Clean up any stray Sheet1
while 'Sheet1' in wb.sheetnames and wb.sheetnames.index('Sheet1') != 0:
    del wb['Sheet1']

wb.save(output_path)

# Verify
wb2 = openpyxl.load_workbook(output_path, data_only=True)
ws2 = wb2.active
print(f"Sheet: {ws2.title}")
print(f"Col I header: {ws2.cell(row=3, column=9).value}")
print(f"Total rows: {ws2.max_row}")
print(f"Mapped: {mapped}")
print(f"Unmapped unique: {len(unmapped)}")
print(f"Coverage: {mapped/(mapped+len(unmapped))*100:.1f}%")
print(f"\nOutput: {output_path}")
