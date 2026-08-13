# -*- coding: utf-8 -*-
"""通用台账统计工作簿生成器（excel-statistics-workbook skill 模板）
用法：修改下方 CONFIG 后运行  python ledger_stats_workbook.py
产出三sheet：统计汇总（分组×分类 项目数/合同额 + 合计）/ 明细（异常行高亮）/ 待补充清单
并在输出同目录生成 <out>.json 统计结果（供子 agent 独立复核对比用）
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONFIG = dict(
    src=r'C:\path\to\source.xlsx',          # 原始台账
    out=r'C:\path\to\output.xlsx',          # 输出工作簿
    sheet='Sheet1',                          # 源 sheet 名
    header_row=1,                            # 表头行号（1-based）
    entity_col=2,                            # 分组维度列（企业/单位），1-based
    name_col=3,                              # 名称列
    amount_col=4,                            # 金额列
    stage_col=5,                             # 分类列
    note_col=6,                              # 备注列，无则 None
    # 原始值 → 目标分类 映射（务必先对分类列跑 Counter 枚举全部取值再填！）
    stage_map={'结算编制中': '结算资料编制中', '一审中': '一审', '二审中': '二审',
               '结算仲裁': '仲裁', '已完成结算': '已完成结算'},
    amount_unit='万元',                      # 金额单位，写进标题
    title='台账统计',
)

thin = Side(style='thin', color='999999')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_FILL = PatternFill('solid', fgColor='D9E1F2')
TOTAL_FILL = PatternFill('solid', fgColor='FFF2CC')
WARN_FILL = PatternFill('solid', fgColor='FFE699')
HDR_FONT = Font(bold=True)
CTR = Alignment(horizontal='center', vertical='center')
RGT = Alignment(horizontal='right', vertical='center')
LFT = Alignment(horizontal='left', vertical='center')


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook(cfg):
    cats = list(dict.fromkeys(cfg['stage_map'].values()))
    wb_src = openpyxl.load_workbook(cfg['src'], data_only=True)
    ws_src = wb_src[cfg['sheet']]

    rows = []
    unmapped_seen = set()
    for row in ws_src.iter_rows(min_row=cfg['header_row'] + 1, values_only=True):
        ent = str(row[cfg['entity_col'] - 1]).strip() if row[cfg['entity_col'] - 1] is not None else ''
        if not ent:
            continue
        raw_amt = row[cfg['amount_col'] - 1]
        amt = raw_amt if isinstance(raw_amt, (int, float)) else None
        stage_raw = str(row[cfg['stage_col'] - 1]).strip() if row[cfg['stage_col'] - 1] is not None else ''
        cat = cfg['stage_map'].get(stage_raw)
        if cat is None:
            cat = '(未映射:' + stage_raw + ')'
            unmapped_seen.add(stage_raw)
        rows.append({'ent': ent, 'name': row[cfg['name_col'] - 1], 'amt': amt,
                     'stage_raw': stage_raw, 'cat': cat,
                     'note': row[cfg['note_col'] - 1] if cfg['note_col'] else None})
    if unmapped_seen:
        print('警告：分类列存在映射外取值 →', sorted(unmapped_seen))
        raise SystemExit('先补全 CONFIG["stage_map"] 再跑，避免统计口径错。')

    # ---- 统计 ----
    stats = {}
    for r in rows:
        stats.setdefault(r['ent'], {c: {'count': 0, 'amount': 0.0} for c in cats})
        stats[r['ent']][r['cat']]['count'] += 1
        if r['amt'] is not None:
            stats[r['ent']][r['cat']]['amount'] += r['amt']
    totals = {c: {'count': 0, 'amount': 0.0} for c in cats}
    for r in rows:
        totals[r['cat']]['count'] += 1
        if r['amt'] is not None:
            totals[r['cat']]['amount'] += r['amt']
    missing = [r for r in rows if r['amt'] is None]

    # 结果 JSON（供子 agent 独立复核对比）
    result = {e: {c: {'count': v[c]['count'], 'amount': round(v[c]['amount'], 2)} for c in cats}
              for e, v in stats.items()}
    result['__TOTAL__'] = {c: {'count': v['count'], 'amount': round(v['amount'], 2)} for c, v in totals.items()}
    result['__GRAND__'] = {'count': len(rows),
                           'amount': round(sum(v['amount'] for v in totals.values()), 2)}
    json_path = os.path.splitext(cfg['out'])[0] + '.json'
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    # ---- Sheet1 统计汇总 ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '统计汇总'
    ws.merge_cells('A1:%s1' % get_column_letter(2 + 2 * len(cats)))
    c = ws['A1']; c.value = '%s（单位：%s）' % (cfg['title'], cfg['amount_unit'])
    c.font = Font(bold=True, size=14); c.alignment = CTR
    headers = ['企业']
    for cat in cats:
        headers += ['%s\n项目数' % cat, '%s\n合同额' % cat]
    headers += ['合计\n项目数', '合计\n合同额']
    ws.append(headers)
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=j)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    r_idx = 3
    for ent in stats:
        row_vals = [ent]
        for cat in cats:
            row_vals += [stats[ent][cat]['count'], round(stats[ent][cat]['amount'], 2)]
        row_vals += [sum(stats[ent][c]['count'] for c in cats),
                     round(sum(stats[ent][c]['amount'] for c in cats), 2)]
        ws.append(row_vals)
        for j in range(1, len(row_vals) + 1):
            cell = ws.cell(row=r_idx, column=j)
            cell.border = BORDER
            if j == 1:
                cell.font = Font(bold=True)
            elif j % 2 == 0:
                cell.alignment = CTR
            else:
                cell.alignment = RGT
                cell.number_format = '#,##0.00'
        r_idx += 1
    row_vals = ['合计']
    for cat in cats:
        row_vals += [totals[cat]['count'], round(totals[cat]['amount'], 2)]
    row_vals += [len(rows), round(sum(v['amount'] for v in totals.values()), 2)]
    ws.append(row_vals)
    for j in range(1, len(row_vals) + 1):
        cell = ws.cell(row=r_idx, column=j)
        cell.fill = TOTAL_FILL; cell.font = Font(bold=True); cell.border = BORDER
        cell.alignment = RGT if j > 1 and j % 2 == 1 else CTR
        if j > 1 and j % 2 == 1:
            cell.number_format = '#,##0.00'
    r_idx += 2
    if missing:
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=len(headers))
        ws.cell(row=r_idx, column=1).value = (
            '注：%d 行合同额缺失（见"待补充清单"sheet），其项目数已计入统计、金额按 0 计。' % len(missing))
        ws.cell(row=r_idx, column=1).font = Font(size=9, color='C00000')
    ws.row_dimensions[2].height = 32
    set_widths(ws, [10] + [9, 13] * (len(cats) + 1))

    # ---- Sheet2 明细 ----
    ws2 = wb.create_sheet('明细')
    ws2.append(['序号', '企业名称', '名称', '金额(%s)' % cfg['amount_unit'], '原始分类', '映射分类', '备注', '异常标记'])
    for cell in ws2[1]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER; cell.alignment = CTR
    for i, r in enumerate(rows, 1):
        bad = r['amt'] is None
        ws2.append([i, r['ent'], r['name'], r['amt'] if r['amt'] is not None else '',
                    r['stage_raw'], r['cat'], r['note'] or '', '待补充金额' if bad else ''])
        rr = ws2.max_row
        for j in range(1, 9):
            cell = ws2.cell(row=rr, column=j)
            cell.border = BORDER
            if bad:
                cell.fill = WARN_FILL
            if j == 4:
                cell.number_format = '#,##0.00'; cell.alignment = RGT
            elif j in (1, 5, 6, 8):
                cell.alignment = CTR
    ws2.freeze_panes = 'A2'
    set_widths(ws2, [6, 12, 60, 14, 12, 14, 12, 12])

    # ---- Sheet3 待补充清单 ----
    ws3 = wb.create_sheet('待补充清单')
    ws3.append(['序号', '企业名称', '名称', '原始分类', '映射分类', '备注'])
    for cell in ws3[1]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER; cell.alignment = CTR
    for i, r in enumerate(rows, 1):
        if r['amt'] is None:
            ws3.append([i, r['ent'], r['name'], r['stage_raw'], r['cat'], r['note'] or ''])
            rr = ws3.max_row
            for j in range(1, 7):
                cell = ws3.cell(row=rr, column=j)
                cell.border = BORDER; cell.fill = WARN_FILL
                if j in (1, 4, 5):
                    cell.alignment = CTR
    set_widths(ws3, [6, 12, 60, 12, 14, 12])

    wb.save(cfg['out'])
    return len(rows), len(missing), json_path


if __name__ == '__main__':
    n, m, jp = build_workbook(CONFIG)
    print('已生成:', CONFIG['out'])
    print('明细行数:', n, '| 缺金额行:', m)
    print('结果JSON:', jp)
